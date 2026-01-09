import os
import re
import pandas as pd
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import glob
import logging

def setup_logging():
    """设置日志"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    return logging.getLogger("statistics_output")

def parse_evaluation_file(file_path):
    """
    解析评估文件，提取各种指标
    返回字典包含所有指标
    """
    metrics = {}
    
    if not os.path.exists(file_path):
        return None
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 解析各种指标
        patterns = {
            'top1': r'Docking Score - Top 1: ([\d\.-]+|N/A)',
            'top10': r'Docking Score - Top 10 Mean: ([\d\.-]+|N/A)', 
            'top100': r'Docking Score - Top 100 Mean: ([\d\.-]+|N/A)',
            'novelty': r'Novelty \([^)]+\): ([\d\.]+)',
            'diversity': r'Diversity \(Top 100\): ([\d\.]+)',
            'qed': r'QED - [^:]+: ([\d\.]+|N/A)',
            'sa': r'SA Score - [^:]+: ([\d\.]+|N/A)'
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, content)
            if match:
                value_str = match.group(1)
                if value_str == 'N/A':
                    metrics[key] = None
                else:
                    try:
                        metrics[key] = float(value_str)
                    except ValueError:
                        metrics[key] = None
            else:
                metrics[key] = None
                
    except Exception as e:
        print(f"解析文件 {file_path} 时出错: {str(e)}")
        return None
    
    return metrics

def collect_all_statistics(output_dir, logger):
    """
    收集所有受体蛋白的所有代统计数据
    返回字典结构: {target: {generation: metrics}}
    """
    logger.info(f"开始收集统计数据，输出目录: {output_dir}")
    
    all_stats = {}
    
    # 查找所有受体蛋白目录 (直接在输出目录下的子目录)
    target_dirs = [d for d in glob.glob(os.path.join(output_dir, '*')) if os.path.isdir(d)]
    
    if not target_dirs:
        logger.warning(f"在目录 {output_dir} 下未找到任何受体蛋白目录。")

    for target_dir in sorted(target_dirs):
        target_name = os.path.basename(target_dir)
        logger.info(f"处理受体: {target_name}")
        
        all_stats[target_name] = {}
        
        # 确定 generations 目录路径，以兼容两种结构
        generations_base_path = os.path.join(target_dir, "generations")
        if not os.path.isdir(generations_base_path):
            logger.info(f"在受体 {target_name} 中未找到 'generations' 目录，将在受体主目录中查找 'generation_*' 目录。")
            generations_base_path = target_dir

        # 查找所有代目录
        generation_dirs = glob.glob(os.path.join(generations_base_path, "generation_*"))
        
        if not generation_dirs:
            logger.warning(f"在路径 {generations_base_path} 中未找到 'generation_*' 目录, 跳过受体 {target_name}。")
            continue
        
        for gen_dir in sorted(generation_dirs, key=lambda d: int(re.search(r'generation_(\d+)', d).group(1))):
            try:
                gen_num_match = re.search(r'generation_(\d+)', os.path.basename(gen_dir))
                if not gen_num_match:
                    logger.warning(f"无法从目录名 {gen_dir} 中解析代数, 跳过。")
                    continue
                gen_num = int(gen_num_match.group(1))
            except (ValueError, AttributeError):
                logger.warning(f"无法从目录名 {gen_dir} 中解析代数, 跳过。")
                continue

            # 根据用户要求，从第一代开始统计
            if gen_num < 1:
                logger.info(f"  - 跳过第 {gen_num} 代 (仅统计 >= 1 代).")
                continue
            
            # 查找评估文件 (适应新的文件名格式)
            eval_file_pattern = os.path.join(gen_dir, f"generation_{gen_num}_evaluation*.txt")
            eval_files = glob.glob(eval_file_pattern)

            if eval_files:
                eval_file = eval_files[0] # 使用找到的第一个匹配文件
                if len(eval_files) > 1:
                    logger.warning(f"  - 找到多个评估文件，使用第一个: {eval_file}")

                metrics = parse_evaluation_file(eval_file)
                if metrics:
                    all_stats[target_name][gen_num] = metrics
                    logger.info(f"  - 第{gen_num}代数据收集完成")
                else:
                    logger.warning(f"  - 第{gen_num}代评估文件解析失败: {eval_file}")
            else:
                logger.warning(f"  - 第{gen_num}代评估文件不存在，匹配模式: {eval_file_pattern}")
    
    logger.info(f"数据收集完成，共处理 {len(all_stats)} 个受体蛋白")
    return all_stats

def create_excel_statistics(all_stats, output_file, logger):
    """
    创建Excel统计表格
    """
    logger.info(f"开始创建Excel表格: {output_file}")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "All Statistics"
    
    # 设置表头
    headers = ['10服务器', '', 'top100', 'top10', 'top1', 'Nov', 'Div', 'QED', 'SA']
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    current_row = 2
    
    # 为每个受体蛋白添加数据
    for target_name in sorted(all_stats.keys()):
        target_stats = all_stats[target_name]
        
        # 为每一代添加数据
        for gen_num in sorted(target_stats.keys()):
            metrics = target_stats[gen_num]
            
            # 填写数据
            ws.cell(row=current_row, column=1, value=target_name)
            ws.cell(row=current_row, column=2, value=f"generation {gen_num}")
            ws.cell(row=current_row, column=3, value=metrics.get('top100'))
            ws.cell(row=current_row, column=4, value=metrics.get('top10'))
            ws.cell(row=current_row, column=5, value=metrics.get('top1'))
            ws.cell(row=current_row, column=6, value=metrics.get('novelty'))
            ws.cell(row=current_row, column=7, value=metrics.get('diversity'))
            ws.cell(row=current_row, column=8, value=metrics.get('qed'))
            ws.cell(row=current_row, column=9, value=metrics.get('sa'))
            
            current_row += 1
    
    # 计算最后一代的平均值
    last_gen_metrics = calculate_last_generation_averages(all_stats, logger)
    
    if last_gen_metrics:
        # 添加平均值行
        ws.cell(row=current_row, column=1, value="计算均值")
        ws.cell(row=current_row, column=2, value="")
        ws.cell(row=current_row, column=3, value=last_gen_metrics.get('top100'))
        ws.cell(row=current_row, column=4, value=last_gen_metrics.get('top10'))
        ws.cell(row=current_row, column=5, value=last_gen_metrics.get('top1'))
        ws.cell(row=current_row, column=6, value=last_gen_metrics.get('novelty'))
        ws.cell(row=current_row, column=7, value=last_gen_metrics.get('diversity'))
        ws.cell(row=current_row, column=8, value=last_gen_metrics.get('qed'))
        ws.cell(row=current_row, column=9, value=last_gen_metrics.get('sa'))
        
        # 设置黄色背景
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for col in range(1, 10):
            ws.cell(row=current_row, column=col).fill = yellow_fill
    
    # 保存文件
    try:
        wb.save(output_file)
        logger.info(f"Excel文件保存成功: {output_file}")
    except Exception as e:
        logger.error(f"保存Excel文件失败: {str(e)}")
        raise

def calculate_last_generation_averages(all_stats, logger):
    """
    计算所有受体蛋白最后一代的平均值
    """
    logger.info("计算最后一代平均值")
    
    last_gen_data = []
    
    # 收集每个受体蛋白的最后一代数据
    for target_name, target_stats in all_stats.items():
        if target_stats:
            # 获取最大代数
            max_gen = max(target_stats.keys())
            last_gen_metrics = target_stats[max_gen]
            last_gen_data.append(last_gen_metrics)
            logger.info(f"  - {target_name}: 第{max_gen}代")
    
    if not last_gen_data:
        logger.warning("没有找到最后一代数据")
        return None
    
    # 计算平均值
    averages = {}
    metrics_to_average = ['top100', 'top10', 'top1', 'novelty', 'diversity', 'qed', 'sa']
    
    for metric in metrics_to_average:
        values = []
        for data in last_gen_data:
            value = data.get(metric)
            if value is not None:
                values.append(value)
        
        if values:
            averages[metric] = sum(values) / len(values)
            logger.info(f"  - {metric}: {averages[metric]:.4f} (来自{len(values)}个受体)")
        else:
            averages[metric] = None
            logger.warning(f"  - {metric}: 无有效数据")
    
    return averages

def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='统计多受体进化实验结果')
    parser.add_argument('--output_dir', type=str, required=True,
                        help='进化实验的输出目录')
    parser.add_argument('--excel_output', type=str, 
                        default='all_statistics.xlsx',
                        help='输出Excel文件路径')
    
    args = parser.parse_args()
    
    logger = setup_logging()
    
    if not os.path.exists(args.output_dir):
        logger.error(f"输出目录不存在: {args.output_dir}")
        return
    
    try:
        # 收集所有统计数据
        all_stats = collect_all_statistics(args.output_dir, logger)
        
        if not all_stats:
            logger.error("未找到任何统计数据")
            return
        
        # 创建Excel表格
        create_excel_statistics(all_stats, args.excel_output, logger)
        
        logger.info("统计输出完成!")
        
    except Exception as e:
        logger.error(f"统计过程中发生错误: {str(e)}")
        raise

if __name__ == "__main__":
    main()